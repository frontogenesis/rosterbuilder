# Import modules
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from subprocess import call

# Define filenames
hr_fname = 'work.xls'
hleos_fname = 'roster.xlsx'

# Convert the HR file to xlsx readable
# If HR changes to xlsx, this code should still support it
print 'Opening Excel Speadsheet...'
xlsBook = xlrd.open_workbook(hr_fname)
workbook = Workbook()

for i in xrange(0, xlsBook.nsheets):
    xlsSheet = xlsBook.sheet_by_index(i)
    sheet1 = workbook.active if i == 0 else workbook.create_sheet()
    sheet1.title = 'ROSTER'

    for row in xrange(0, xlsSheet.nrows):
        for col in xrange(0, xlsSheet.ncols):
            sheet1.cell(row=row+1, column=col+1).value = xlsSheet. \
            cell_value(row, col)

# Hide rows and columns from HR that we don't need
print 'Formatting Spreadsheet...'
sheet1.column_dimensions['E'].hidden = True
sheet1.column_dimensions['G'].hidden = True
sheet1.column_dimensions.group('J','R', hidden=True)
sheet1.column_dimensions.group('U','W', hidden=True)
sheet1.row_dimensions[1].hidden = True

# Adjust column width so it's easier to read data
sheet1.column_dimensions['A'].width = 10.0
sheet1.column_dimensions['B'].width = 20.0
sheet1.column_dimensions['C'].width = 15.0
sheet1.column_dimensions['D'].width = 12.0
sheet1.column_dimensions['F'].width = 30.0
sheet1.column_dimensions['H'].width = 30.0
sheet1.column_dimensions['I'].width = 80.0
sheet1.column_dimensions['S'].width = 25.0
sheet1.column_dimensions['T'].width = 30.0

# Bold headers
maxcol = sheet1.max_column
for row in range(2,3):
    for col in range(1,maxcol):
        cell = sheet1.cell(row=row,column=col)
        bolded = cell.font.copy(bold=True)
        cell.font = bolded
        filled = PatternFill(start_color='00bfbfbf', end_color='00bfbfbf', \
        fill_type='solid')
        cell.fill = filled

# Generate new sheets inside of the workbook
sheet2 = workbook.create_sheet(title='PROCEDURE')
sheet3 = workbook.create_sheet(title='Data')

# Determine NEO date
#neodate = sheet1['E4'].value
#print neodate

# Store courses in a dictionary
courses = {'01':'Health Link Outpatient Clinicians (Nurse/MA/Therapy/Techs)',
'02':'Health Link Outpatient Clinicians (Nurse/MA/Therapy/Techs) - Day 1 Only',
'03':'Health Link Outpatient Clinicians (Nurse/MA/Therapy/Techs) - Day 1 with \
Day 2 Scheduling',
'04':'Health Link Outpatient Clinicians (Nurse/MA/Therapy/Tech) - Day 2 Only \
- Orders',
'05':'Health Link Outpatient Clinicians (Nurse/MA/Therapy/Tech) - Days 1 & 2',
'06':'Health Link for Receptionist/Schedulers (Day 1 & 2)',
'07':'Health Link for Receptionist/Schedulers - Day 1 Only',
'08':'Inpatient Basics',
'09':'Inpatient Basics; Inpatient 101 - RN/RT; Inpatient 102 - RN/RT',
'10':'Inpatient Basics; Inpatient Workshop',
'11':'Inpatient Basics; Inpatient HUC',
'12':'Radiology - Technologist',
'13':'Emergency Department RN',
'14':'None',
'15':'TSC/MSC Surgical Services Basics',
'16':'Surgical Services Basics'}

# Enter trainees from calendar, LDS, or other sources
enterdata = raw_input("Do you need to enter a trainee manually? ")
while True:
    if enterdata == 'yes' or enterdata == 'y':
        source    = raw_input("Enter source (i.e. LDS, Calendar): ")
        lastname  = raw_input("Enter trainee's last name: ")
        firstname = raw_input("Enter trainee's first name: ")
        jobtitle  = raw_input("Enter job title: ")
        dept      = raw_input("Enter department name/location: ")
        options   = courses.keys()
        options.sort()
        for entry in options:
            print entry,courses[entry]
        course    = raw_input("Enter class number: ")
        course = courses[course]

        # add new row, calculate cell position, add data
        newrow = sheet1.max_row + 1
        sourcecell = sheet1.cell(row=newrow,column=1)
        sheet1[sourcecell.coordinate] = source
        lastnamecell = sheet1.cell(row=newrow,column=2)
        sheet1[lastnamecell.coordinate] = lastname
        firstnamecell = sheet1.cell(row=newrow,column=3)
        sheet1[firstnamecell.coordinate] = firstname
        jobtitlecell = sheet1.cell(row=newrow,column=6)
        sheet1[jobtitlecell.coordinate] = jobtitle
        deptcell = sheet1.cell(row=newrow,column=8)
        sheet1[deptcell.coordinate] = dept
        coursecell = sheet1.cell(row=newrow,column=9)
        sheet1[coursecell.coordinate] = course
        print lastname,firstname,'has been added to the roster'
        # ask user to add another trainee
        another   = raw_input("Do you need to add another trainee? ")
        if another == 'yes' or another == 'y': continue
        else : break
    else:
        print 'OK.  I just wanted to make sure.'
        break

# Data validation for cells that do not have classes assigned
dv = DataValidation(type='list', formula1='"Test1,Test2,Test3"',
     allow_blank=True)

# Data validation error messages
dv.error = 'Your entry is not in the list.'
dv.errorTitle = 'Invalid Entry'

# Data validation prompt messages
dv.prompt = 'Please select from the list.'
dv.promptTitle = 'List Selection'

# Add data validation object to roster worksheet
sheet1.add_data_validation(dv)

# Apply validation to the range of cells requiring them
# Need cells I3 to the last row in the I column
maxrow = sheet1.max_row
maxvalidationcell = sheet1.cell(row=maxrow,column=9)
maxvalidationcell = sheet1[maxvalidationcell.coordinate]
dv.ranges.append('I3:maxvalidationcell')

# store output file on U drive with following convention
# with following format:  'Wk of [Monday following current day] Anticipated
#HL Training Attendance.xlsx'# Generate new sheet inside of the workbook
print 'Saving Spreadsheet...'
workbook.save(filename=hleos_fname)
print 'Done!'

#NEO date.  Column E.

# go to network drive
# U:\UWHealth\HLDoc\Training\SharedUW\NEO (New Employee Orientation)\
# NEO Rosters\2016
# open "Standard Template.xls"

# copy/paste from e-mail version to network drive version

#outinp = raw_input("Is this person working in an inpatient,\
#outpatient role, or both roles? ")

################## #
# CBT in the role  #
# Is a CBT needed? #
# cbt =            #
####################


########################
# Send an e-mail out   #
########################

# send to IS - Systems Security - UWH
# send to HR - Learning & Development General Inquiries-UWHC
# send to IS - Health Link Education-UWH

# text = str("...body of email...")

viewxls = raw_input('Would you like to view the spreadsheet? ')
if viewxls == 'y' or viewxls == 'yes':
    call(["open", hleos_fname])
    print 'Your spreadsheet should appear. Looking forward to your next visit!'
else:
    print 'The program has completed. Looking forward to your next visit!'
